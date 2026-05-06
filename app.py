from flask import Flask, render_template, request, jsonify, redirect, url_for, flash
import json, os, random, secrets
from datetime import datetime, timedelta
import pandas as pd
from werkzeug.utils import secure_filename
from db import get_db, get_db_schema
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
import bcrypt
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login_view'
login_manager.login_message = 'Debes iniciar sesion para acceder.'

class User(UserMixin):
    def __init__(self, id, username, rol):
        self.id = id
        self.username = username
        self.rol = rol

@login_manager.user_loader
def load_user(user_id):
    with get_db() as db:
        row = db.fetchone('SELECT * FROM usuarios WHERE id=?', (int(user_id),))
        if row:
            return User(row['id'], row['username'], row['rol'])
    return None

def rol_required(*roles):
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated(*args, **kwargs):
            if current_user.rol not in roles:
                return render_template('acceso_denegado.html'), 403
            return f(*args, **kwargs)
        return decorated
    return decorator
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
MHE_LOGO_B64 = 'data:image/png;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCACqAOQDASIAAhEBAxEB/8QAHQABAAIDAQEBAQAAAAAAAAAAAAYHBAUIAgMBCf/EAEUQAAAGAQIDBQQGCAQEBwAAAAABAgMEBQYHERIhMQgTQVFhFBUicSMyVYGT0RgkN0KRobHBUnN0gxYXNGIlM0NTcqLx/8QAGgEBAAMBAQEAAAAAAAAAAAAAAAECAwQFBv/EADURAAEEAAQCBwcEAgMAAAAAAAEAAgMRBBIhMRNBBRRRYXGR0SIygaGxwfAWU1SSUvFCYuH/2gAMAwEAAhEDEQA/AOywAARAAARAAARAAARAAARAAARB+KUSUmpR7ERbmY/RTPaE1trtPTRRw4pz7mQ3xcBKIksp81H/AGGsML5nhjBZWckjY25nFXI04h1sltqJST6GQ9Dn7s/a/QsutGsWvIZwbJwjOM7xEaHj/wAPoY6BEzwPgfkeKURStlbmaUAAGK1QAAEQAAEQAAEQAAEQAAEQAAEQAAEQAAEQAAEQAAEQAAEQAAEXl0lm0smjIl7HwmfQj8Bydqti0Ww1TmRMm92Srhdeucp5DKyT3TZH8PXrsQ6yUokpNSjIiItzM/Ac36pMe8NbpVrDlQnYR45Ijk6UlO3emkyJHXqPQ6PcWvNaaLjxjQWjxUCwPGKVWV4i9RtV0axt2ly691bS9muA9j4ufoOx69MtMJlM5xpySSC71TZbJNXjsR+A5U00rn4GWaXSZjsNlmtrn0TFqkp2ZUajMiPn4jq9l1t5pLrLiXG1lulST3Iy8yMX6ScS4a3v9Sq4EU0/nJewAB5i7kAABEAABEAABEAABEAABEAABEAABEAABEAABEAABEAAPkW5giCDWeoMWTkqsWxVkri2b/6lSD+giF5rV5+hcxVesmqVzlWYt6V6aP8A65IX3U+xb5kyn94kn6F1MY+pmV0mgeBs4biPA9kkxvjfkq+JZGf1nVn4mfPYh6EWCd7IItzth3dpXHJiRqQdBufsFa+S5/XUNhDxZSyu8lnHwpgxi34S8VK/wpL1Guv3NO6PIaXHJ9NBXc3C/gjNJI+HxNR+RCsOz3XM4Rp1c6yZg4uRZz21ONOPnus0eBFv0NR/yFUaTZLZZf2k6nIbZ5TkiVKUpJGfJtG3JJehEOpmCFvynRo1PafQLB+J93MNXbdwXU6HNNnNQJOCv08Nm1SwT7aXEFwvpPy9S8hnVme1MbJF4HZoLH7ZtH6klexNSG/A2z6fcOY+1zYTanXli0rX1MTIsdtxpxJ7GRlsLIzdqNrloJGy2tSTeS0ye83bPZaHEfXT6Ef1iFXYQZI3vJyu+R9FZuI9pzWjUfMK2JOfJxu/ZpczZTXpkq4YdiX/AE7x+CTP91XzE7QpK0JWhRKSotyMj3IyHLOg+pFbqZjr2l+o6W5MlxrgiyHfrO7dC38Fl4GMrDM6vNGNQP8AlznUp2Zjzyv/AAuyd3NTaDP4SUfiRdD8hjLgXWWAe0OXaO0ei0jxQoOPun5eK6dAeWXG3mkOtLSttaSUlST3IyPoZD0PNXagAAIgAAIgAAIgAAIgAAIgAAIgAAIgAAIgAAIgpTtX6mqwjDyqKt4k3NqlTbZkfNpv95fz8hdTikoQpaj2SkjMz9BwBqZaStVNfFRmFqcZfnJgxSLmSW0n1/qPS6Mw4llzP91upXFjZjHHTdyrZ7ONRB030ottVshSRzJbalRu8+saPAi9VGKDiLtdTdVI5z3Fuy7icklmf7jZq6F6EQurtmXbdVV47p1Wn3UaMyl59CT5GRFskv5GYinYzp0WWsBTHEcSK+Ktwty/eMtiHsQuLYX4t2528OS82UZpWYcbDfx5qY9tC+aqqjH9Oq1RNsMspefQjkRpSWySP7y3FS9nH9tmOf5yv6D97R92d7rNeySWammHSjteiUkX9x+dnH9tmOf5yv6DaKLh4IjnRPmsnyZ8WD3qUds79sqv9GgbHsW5V7szqZikpe8O4ZM0IM+XeEX9y3Gu7Z37ZVf6NArLT21co88o7VpRoUxMbMzLyMyI/wCoiOIS4EM7kkkMeLLu9SPWuilYDrJZM1ylRlNSSmQ1p5cJKPiLYXrmzUXXXs9NZJFbQeQ06OJxKS+LjSXxp+R9SEb7c9Y178x3ImC5S45tqPbr0NP8hpexhlHuzP5WLyV/qVyyZEgz5d4Rcv4luMHEy4VmIb7zftutm1HiHQn3XfgVi9jTU523rl4LdyDXNhI4oS1n8S2i6p+ZDpMfz+zRqVpP2g35ELibRDnlIZIuRKZWfT5DvamnsWlTEsYyiUzJaS6gy8jLceX0nA1rxKzZ2q9DAylzTG7dqywAB5a7kAABEAABEAABEAABEAABEAABEAABEAABFFtW7U6TTW/tErNCmISzSZeZlt/ccc9j6q97a0xZTyeMobLkkzPwX4GOsta5GMS8UkYxktq7XNWjZoJbaTNWxHzFG4o/p3otZqucacu8qkzW+5W202f0KS8T5D2cE7LhnsaDmdt/tebigHTNcToFU3aWtlXGtd87x8bcdwo7Z/8AakWh2E2EJscqsTT8TUdCSP7zMfCZp/pJk8x/I5+WW8GVZLOQ7GW2fE0pXVJ8vAbXFL3BNGDlxcbbvcoTbpIn3Gmz2YIt/T1HdNKH4bgMBugNuxckUeWfiuIpc25ZIVMyu3lK5m7NdP8A+xiX9nH9tmOf5yv6CyP+VWj0xaprubWjSpCjeU2bZ7oNR7mR8vDcZlNi2lWnlrHzGov7i7m1yuJqC22fE6Z8vIdEmLY+IxtBsitlizDObIHkivFRPtnftlV/o0ClkrNtaHC6oWlRfce46cyaJprq/bKyzIbC6xqaSCYOG62e5kXRRchrFaQ6Mmky/wCObPp/7Z/kK4fFMhibG8GwOxWnw5klL2kV4rfdq9JTtCsNtlc17slv82xzvpvaOUuoNDaNqNKmJrZ7+hnt/cdD3eTYNqDQQtOLeNf01XUmn2e0W2fC93ZcJeHiXMaROlGjUYykpza1WpkycSlKD3UaeexcvQZYaZsMJikB1vlyKvPEZJA9pHJfLtz1aUZfSXaElwzYhoNRePD0/qL37LFw5caK0rjqzWuMlUYzP/t//RVGU5BgWsaIdXkse+xtuoI/Z5LjZ7PlsReXoLU0BRhOOU54fi15IstnFyfpknuW+244cST1RsTgczfp4rrgA6wXgiirUAAHir00AABEAABEAABEAABEAABEAABEAABEAABFpsitsarVtJv51ZGUsj7opa0EZ/LiFIdpK5kT6irTptkdIw+l5Ryu5kNJM08tvD5i0dS9LcS1CkQ38kiOPrhpNLRpcNOxH1ERLs1aXF0rJBf75jvwr4IyHuJscqsfVck7ZX20AUt7gmSYkzh1U1kF7jzlqmMkpSjdaMzX47iru0XbW8+3qFac5NSsRUEftRMyGk7nv48hNP0adLfsuR+OYF2atLi6Vkkv98xrHLho5OJZPwFfVUfHM5mWh5qW0+T4ImphpnXeOqlEwgnj71rmvhLf+Yjur2Q0UrT6zYw6/oGrpSCKOpDzRK33LfY/kMP9GnS37LkfjmP0uzTpaXSrkfjmKNOFa4OzHyHqrOExbloLXdnq8Yg4StnUPIaJ+078zQbr7SlEjbzFjrynTzgVw3WOkrbl9K11EJPs1aXGe51kg/8AfMfn6NOlv2XI/HMJHYWR5dmIvuHqjGztaBQUD0mtcgjauWcrLcnpnMbWbvcIckNGjmr4Ni28hd1hlGAqgSCi3eOk+bSu6PvWuStj2/mIcfZq0uMtjrJJl/nmPz9GnS37LkfjmLzS4WV2ayPAD1VY45mNqh5qE9ny2uoOT27moWTUr1etP6ql2Q0ot9/Dly5C/qG7xGwmmxSWNRIlcJmaIy0Gvb7uewrg+zVpcfWskn/vmJDp7o3hGC353dBBdZmG0bXEpw1Fwn1FMTJh5SXAm+yqH1VoGSxgNIFKxAAB5y7EAABEAABEAABEAABEAB4kOEyw46ZGokJNRkXU9gRewEJotTMctpUFlCpEdNg4tqI482aUvLR9ZJH5jOyDNq2myFqhdjS35zrBvttst8XEguo0MLwarVZ8VhF2pQAi9/m9XSVlbOnMSklYuE0w2Tfx8Z9CMvAfNvUChXj9rcEp4k1KuGcwaDJxk/IyDhPIulPEZdWpYAh1dlK2/Yp85anYl46gqxDbfNtJp3+P59R5udR6Kpt7CulolEqtJCpjhNmaWkr+qo/QxPBeTQCjitAslfXMradHtW4LEwoDSY6nzc2LidNP7iTPkMGVfTZzkRpm1TVsKhFJS+6kt3leKefLkNlk+Q46zY0tZZRval2y9oZ93xJM+vXwH0z21x/HqePMu4KXYpPoZaJLJK4VKPYiIvAaN/4jLqVR3M5lFb7K8gKJXvx1dwtcNTy0lwkSjI+vxeB+g2d7k1o3Wznoq0pNEaOts+HYyNZ7K6jZ5VcUEH2NEmt9vmrb448VtolOEnbrt4EM2jsKy/olWCq9TDKyNLrUlnhURJ8DI/AhJcAA7LooANkZlGLi8uYdDUrRLkrckur79bSEOOEkk77ERch5TlOQtMU8ru25DJx3XpqCT8ZoIzIjL1LxISbB7bHcgpG7LHe5XES6tCTSnbhUR7K+Qx8tySjxR6uZmxFG5PdUxGS00R8Supp9ABs5MmqVpmzaLToyKfaRq1Ldq1WNSY65HtS0EXHsrYkc+RcuY3lTcTZWFO2bqUpkobcJKiLkvh3Il7eR9R8aW7xXI2JcMmGUrrT3kxX2iSpjx328vUhgOajY3HVGakMPx6uS57OzMUztHUroRehH59BBaXaBqBwGpcsOLkVvHhPkVk1ZKVWnL7xKS+gXy+E9uXj4jf6eTZ8+nKTYSVPOrSlWxmg+Hf8A+P8AcfZydjVVcxceNEWPJsm1LaaJBETxF1L1HqsscfjZBKx2uJpucwyl55ltO3Ck+m4hxtppvepaKOpXnHr1MlFn7bJZJyJJdQSCMiUSE9OQ1OB5XLt7mZEnoU2l7d6CRtGn6PxIzPqZDxByvE5NPd5MiCaGax1bM1w2fiNST+P57DYuZbj51tNbQ9psSxcSzEfYSSiI1dC9BJZuMqB2xzKUANDHyqufzGTiyUvFOjsE+5un4SQfQ9xrX9QKxNiiNFgz5kdTxMHLYaNTSV77dfL1GQieeS0MjRzUwAaXKsnpsZZiPXEoo6JkhMdkzLqtR7EQZjk9TidN74unjYgk4lC3SLckcR7EZ+gqGONUN1Je0XZ2W6AaORk9c1a1cBBrfOzQa4zrRcTZltvvv8h6tcop6vIq2gmSSbn2XF7M3t9fbqJ4buxM7e1boAAUVkAABEHylqNMV1RJNRkgzJJdT5dB9QBFQOJ4vk9VBp8hdqZcxddPe7+pf23Qhaj2ea9SLmJLn9dZWGplTaN11t7uYr1pdfiHwqJauZJ/mLZAdRxbi7MR2/Nc4wwDcoKqnUyLf3WO4uqFUzmpDNghx1PJTjLaeXErzPxG9sMFabxLJI1a8t20u21LfkPdXHNtk7l4F4CcgKcd1ADSv9q3BbZJ5qs8aK6sY2L1L9JLgLpuD2x17bg3Qjh2T579RG84x23tsyy9ZU9mbU6IwzAdaVs244gv3i8U7+YvABZuJLXZgPy7UOgDm0T+bKpslg5KifgUiTVuzn60ycslxUlwIMk7HsQzNT1XeR4KycTH5iXytGlJjnt3htJPms/IhZoCBiNWmtk4OhF7qsslj5FQ59EzCvqXrevfgpiy4rO3fsmXMlJI+R9eY3GWWF7Z4g01XVMuJKsXUsqS4RcUds/rKV9wmoCvGuiRqFbhb0d1VumtLf4ln9vUuwzco57aZTL7KSJpp4i2Unb1It/mMvWKHayrnE5FbVyZzcCwORJNoi+BHDtuLHATxyZM5Gv4FHBGTJarhGLzr3IclvVsKrE2dUdcwlfJajMv/MVt/AR84WSXGmCNO7LHHmbBpLcY5XI4/AhRbOkfyL+Ji5wEjEkctqr4KDADz/Cqm1HxS2tbWs90MPFZUUEna6arkhT5GXwKPyMtx+YNR3tfqvIu7OueJdlWJXNkEe7aXi/9MvQhbQCesuyZaTgNzZlUeNV91F09zCLIpZRSJVnIejsGRbvNrUWxl9xGPhHwi9ob6jfxdtSMbly2pFjWOHzhuF1W35Fv1IXGAdadZobp1dtDXZV7HgWTuslzKXAkNV8iqRGbl8uE1kZ7kX8Rg6bS8oxaKzhtrjMuT7O8tLFjG2Nlxo1GriVue5GW4tABUz2KI7PkpEVGwe35qrNR8fvMzsrSCiP7PChxDKKt5BGTj/Ulo8jLoPMlN5c6R1cG8x6Q9OJ1qPNiKSRmtCFERrP0NJbi1QEjEEACtlBgBJN7qqMdw7IMa1AgR4LipWHpSt6Olw93IS1Ee7e/ijy8hqc8x/LsjatMkhw1MT6+WhysYWku9NKDIjJJ+BKLmLtASMU4OzVr+fVQcO0ty3osKjlyJ1PFly4q4kh1pKnWF9W1bcyGaADnO66AgAAhEAABEAABEAABEAABFW2rtjqTCsIScHgIksKQZyDU2Stj8BCPf3aA+w2vwS/MdAAOWTDF7ic5C+gwnTkeHhbEcLG6uZBs+Oq5/wDf3aA+w2vwS/MPf3aA+w2vwS/MdAAK9Ud+47zXT+pIv4cX9T6rn/392gPsNr8EvzD392gPsNr8EvzHQAB1R37jvNP1JF/Di/qfVc/+/u0B9htfgl+Ye/u0B9htfgl+Y6AAOqO/cd5p+pIv4cX9T6rn/wB/doD7Da/BL8w9/doD7Da/BL8x0AAdUd+47zT9SRfw4v6n1XP/AL+7QH2G1+CX5iUaX2mrEvKCZzCsRHrO6UfGTZF8fgLYASzDFrgc5KwxPT0c0ToxhY22KsA2O8aoAAOtfOoAACIADGiz4UqU/FjyWnXo5kTyEq3NBn03C1IaSCQNlkgAAoQAAEXxmm8UN446kpdJBmg1FuRHt4iv9Ossv8jxK4tJi4TMmK860wlKfh3R4q+YsCe4TMJ5w0qXwoM+FJbmfLwIUjpvicVeIX8i9pbJib7Q84hB8aTWk/q7EXUcsznB7Q3sK9/ouDDyYSV0u4cyjQJ1JvSxptasDR/JbXKcS99W5R21rdWhKGi2JJJPbcxh4Hn6sgze8oH2SZRGPjgqMtjebLkZ+vPcRHBH7im0WXVIp7FuyceWwSO5PibSs/r/AHEPzKKLI8ZvcUyOA25aOwyKO8zGYMjNgy+I1evUYiZ4Yw9lX8V6z+jMI/E4mM0MxcI9dBl1HdR0Gp5lTZd7kOSWtjDxV6FEjVzncuSZKDX3jm25kRF4F5jW5fkWb4rp45cWKaxyzZkEgyQk+7cQZ7EfoY1NW9kOn2SWbzVFLtsdt3va2lx07vR1qLmlSeozNW5dpkulzyY2Pz0SX30dzHNO6zSRkZmZF0El5LHGzm1WceEjbioWBjDCS3XS/wDsDrfbYOyyMly7Jq25xWtYXAUd4ku8ccQZE0rbfcvQbmwssqq8YvbCRJrZj8VJLh9ynZJ7dUqLzEIzqNKs8gwd4qKwlQoCCOcRNKLgLh22+Y306Q23iV7WUmNWLEVKN0KWlRrecUZHskj57AHut1nw8kkwsPDgysFn3tB/mdzdjShVahZuF6gIzHD5Umv4IN5EaM34j5c0KIuu3ikxqMnzXKqvBKK9aVAVJsJCGXUG2fCklHsRl8thj5RhM2zoYeW4kTlXkrEUm3G1J4CkpItjQsvP1GBncK5f0yxStZpZj85h5p+Q0hs/oySe6t/UVe+UNN71uOf/AKt8Phej3TsdGBlc+i127aBsXzbtR+6k+RZbkeGTq2RkCIs6mmupZXIjoNCo61dNy8SGdq3mzuIUUSbAjlKeeeSZo232Z6rV/Aa/NI8nPamBQRa6VGj+0NPS35DZo7tKOZpIj6mYxl1drlF/dGROQILMX2BlMmOZ943tzWn1MXc5+rWnfb7rkhgwh4UuIaAW3nGwIsBuguibO3IWrJrJjNjWR50ZZKakNJcQoumxluK9ZzC9ptS3ceyx6GxVvMKegSkoNPebdUmfmQaCPXMPH3sbuYEtlVc6pEZ91syS61vyMhi9oWDIsodIxDqZM95meh5ZtNmfA2R/Fz/sLPkc6ISN3HL7LDCYGGLpJ+BlpzHWA7TTmHA/Xy3UvwWVkNjBkzrhyOTTzivYkobNJk3vyUr5iJ4/lWYXlrkVcxKqYiql7um1uoPZzyM+fIWVDebOuaeQ0ttsmyMkGnY0kRdNhSWM1zKcgyqfkGL2jrMqUT0M0IURrJPhy9fMJS5uQAqejmw4jrD3sAIy5RQNe0BoCRem/mrCuciuGZtXjcBEVy9lx++ecPm0ykvrK28efQh5izsux/3jYZZLrZNPGjm6l6O2aF7l4GRiNZREyli9p9RKWlW4tuOceZVqWXedyZ8jL19BvHbCRqFjNpSKorCqakxlNm7KLh4V+BbePPxAPJJ3vl2HRS7DRsiYQ1pjPvnTM05ta1sUNq0I7Vhncah2OMllNS3WpZWg32a5xBmtbXUvj8DMhhWuqcw9PazMKit75KpZR50NRbuJMj2USfUh6wjIr/H8YLGbzG57tlAQbDDjCOJqQkuST4vAYUGhnYnjFFDk178uS9ce3TER2+NLSVHzI/kMi95FtJ217iu5uGwzZC2aNpp/s0feZRvUHbanHWypLcZq5YYVFyPEpUVxLjyG3EPJ3NJqPYyMvAyH0tMmt5OTMYlRnHKxTFKRMlOp3baSfQiT4mZiH51gVpUWzdzhZL922EttdnWkXLrv3iS8D8xs8lrrzF9S2c2rK56yrZkVMaew1zca26KIvEWMkt+13X4doWLMJgHNbwSDYeWh1Xm0pru8a1yOngs+nzG/qM9Zw/L2YyzmoNcCdHLhQ5t1SZeBjY4Zm6bjNL/GJSUNSa9wjY25d434n6mRjVSoknM87p7w66TDq6Qlu8b6DS484ZckknrsQjEukyCRFVlmPVr8fIYNksiQ8g09+wszLn5kW+4cSRp01F+YU9TwU7aeAx5aAdaDX2aPdYAvsu1PGc2KZqq9h0Puzbjwjdec6mThnyIvuGs0jtp1jkWTxnosGO3Cmm2tbKDJb6v8SjGirqqZTayV01NZMWyqCpM6WlszSbyviPn4l4D66dv2VHYZzPfpbAjkSlPwy7k/pi2Mi2+8Q2RxeM3aforTYHDtwz2wAG2MrXXNmon79wUhlagdxqyxiimklAeZ4UyduRv/AODf+IsAURlWLZM/p5Et2VuO2TEwrBqMlgyeJ1R7mgz8uouTFbF61oIk2TEeiSFtl3rLqdlIV4kNoJHFxDvELzOl8Fh44Y5cORpbXV/kOfxHw0WzAAHUvn0AABEAABEAABEAABEAABEAABEAABEAABEAABEAABEAABEAABEAABEAABEAABF//9k='

@app.context_processor
def inject_globals():
    return dict(
        get_logo_b64=lambda: MHE_LOGO_B64,
        current_user=current_user
    )
COOLDOWN_DAYS = 21

PERSONAL = [
    {"id":1,"nombre":"OMAR ABDALA CHAMI ISA","cargo":"COORDINADOR","zona":"Distrito Nacional"},
    {"id":2,"nombre":"JUAN ROSARIO SANCHEZ","cargo":"COORDINADOR","zona":"Distrito Nacional"},
    {"id":3,"nombre":"ROBERTO CARLOS GUZMAN MARTINEZ","cargo":"COORDINADOR","zona":"SD Norte"},
    {"id":4,"nombre":"PEDRO JULIO CASTILLO","cargo":"COORDINADOR","zona":"Distrito Nacional"},
    {"id":5,"nombre":"ERCILIO ANTONIO VASQUEZ GUZMAN","cargo":"COORDINADOR","zona":"SD Oeste"},
    {"id":6,"nombre":"JUAN TOMAS BATISTA BAUTISTA","cargo":"COORDINADOR","zona":"Distrito Nacional"},
    {"id":7,"nombre":"DOMINDO MERCEDES","cargo":"INSPECTOR","zona":"SD Este"},
    {"id":8,"nombre":"CARLOS BENJAMIN HERNANDEZ MENDEZ","cargo":"INSPECTOR","zona":"Distrito Nacional"},
    {"id":9,"nombre":"ELIEZER JIMENEZ DE LA CRUZ","cargo":"INSPECTOR","zona":"Distrito Nacional"},
    {"id":10,"nombre":"JOSE MIGUEL RODRIGUEZ BENITEZ","cargo":"INSPECTOR","zona":"Distrito Nacional"},
    {"id":11,"nombre":"JOSIAS LABOUR VERAS","cargo":"INSPECTOR","zona":"Distrito Nacional"},
    {"id":12,"nombre":"JULIO BIENVENIDO DIAZ SANTANA","cargo":"INSPECTOR","zona":"Distrito Nacional"},
    {"id":14,"nombre":"LEUDIS RAFAEL DIAZ SUAREZ","cargo":"INSPECTOR","zona":"SD Este"},
    {"id":16,"nombre":"PABLO FORTUNATO TAVERAS","cargo":"INSPECTOR","zona":"SD Norte"},
    {"id":18,"nombre":"ROBER JESUS GARCIA","cargo":"INSPECTOR","zona":"SD Este"},
    {"id":19,"nombre":"VICTOR OMAR MERCADO MEJIA","cargo":"INSPECTOR","zona":"San Cristóbal"},
    {"id":20,"nombre":"VLADIMIR TAVERAS MOYA","cargo":"INSPECTOR","zona":"San Cristóbal"},
    {"id":21,"nombre":"WALTERIO RAFAEL PELLERANO CASTILLO","cargo":"INSPECTOR","zona":"Distrito Nacional"},
    {"id":22,"nombre":"FRANKLIN FELIX ALVAREZ","cargo":"INSPECTOR","zona":"Distrito Nacional"},
    {"id":23,"nombre":"MANUEL BOLIVAR PATIN ENCARNACION","cargo":"INSPECTOR","zona":"Distrito Nacional"},
    {"id":24,"nombre":"EDISON ALBERTO MEJIA DE LEON","cargo":"INSPECTOR","zona":"Distrito Nacional"},
    {"id":25,"nombre":"KELMI MOISES DE LA CRUZ BERROA","cargo":"INSPECTOR","zona":"Distrito Nacional"},
    {"id":26,"nombre":"ANGEL DAVID HERNANDEZ CUAS","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":27,"nombre":"DENNY JOSE ACEVEDO MACARIO","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":28,"nombre":"HENRY DE JESUS","cargo":"AUXILIAR","zona":"SD Norte"},
    {"id":29,"nombre":"GABRIEL A FLORENCIO POLANCO","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":30,"nombre":"IVAN MEDINA","cargo":"AUXILIAR","zona":"SD Este"},
    {"id":31,"nombre":"JEISON CAMILO MERCEDES SENA","cargo":"AUXILIAR","zona":"SD Este"},
    {"id":32,"nombre":"JESUS FELIX GOMEZ BONILLA","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":33,"nombre":"JUAN LINARES REYES","cargo":"AUXILIAR","zona":"SD Oeste"},
    {"id":34,"nombre":"LEONEL DAVID HERNANDEZ","cargo":"AUXILIAR","zona":"SD Norte"},
    {"id":35,"nombre":"ANTHONY JUNIOR BRITO","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":36,"nombre":"ELIS ENMANUEL MARTES POZO","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":37,"nombre":"LORENZO WILLIAMS JANSER GIL","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":38,"nombre":"ROBIN DE JESUS GARCIA BELLIARD","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":39,"nombre":"MIGUEL ELIGIO TAVAREZ BELLO","cargo":"AUXILIAR","zona":"San Cristóbal"},
    {"id":40,"nombre":"AURELIO REYES MERCEDES","cargo":"AUXILIAR","zona":"SD Norte"},
    {"id":41,"nombre":"DALBERTO SUAREZ ENRIQUES","cargo":"AUXILIAR","zona":"SD Oeste"},
    {"id":42,"nombre":"CLAUDIO ANTONIO TEJADA CASTRO","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":43,"nombre":"OMAR ANTONIO NUÑEZ AMPARO","cargo":"AUXILIAR","zona":"SD Norte"},
    {"id":44,"nombre":"ANDY ALMANZAR SANCHEZ","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":45,"nombre":"ANDERSON JULIO FELIZ ROMERO","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":46,"nombre":"CRISTIAN ALEXIS LUGO MORILLO","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":47,"nombre":"SIMON DE LOS SANTOS GARCIA RODRIGUEZ","cargo":"AUXILIAR","zona":"SD Oeste"},
    {"id":48,"nombre":"RANDEL RAMON PENA SUERO","cargo":"AUXILIAR","zona":"Distrito Nacional"},
    {"id":49,"nombre":"JOSE LUIS DIAZ ARIAS","cargo":"AUXILIAR","zona":"SD Oeste"},
]

PROV_COORDS = {
    "Distrito Nacional":[18.4861,-69.9312],"SD Este":[18.5001,-69.8500],
    "SD Norte":[18.5850,-69.9500],"SD Oeste":[18.5000,-70.0500],
    "San Cristóbal":[18.4175,-70.1106],"San Cristóbal / Haina":[18.4300,-70.0500],
    "Santiago":[19.4517,-70.6970],"La Romana":[18.4273,-68.9728],
    "Puerto Plata":[19.7930,-70.6880],"San Juan":[18.8053,-71.2284],
    "Espaillat":[19.3941,-70.5232],"Dajabón":[19.5497,-71.7085],
    "San Pedro de Macorís":[18.4586,-69.3058],"La Vega":[19.2211,-70.5294],
}

def init_db():
    with get_db_schema() as db:
        db.executescript('''
            CREATE TABLE IF NOT EXISTS personal_state (
                persona_id INTEGER PRIMARY KEY,
                carga_total INTEGER DEFAULT 0,
                no_disponible INTEGER DEFAULT 0,
                motivo_no_disponible TEXT DEFAULT '',
                motivo_detalle TEXT DEFAULT '',
                conflicto INTEGER DEFAULT 0,
                ultima_asignacion TEXT,
                zona_counts TEXT DEFAULT '{}'
            );
            CREATE TABLE IF NOT EXISTS personal_disponibilidad_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                persona_id INTEGER,
                fecha_registro TEXT,
                motivo TEXT,
                detalle TEXT,
                disponible INTEGER,
                semana TEXT
            );
            CREATE TABLE IF NOT EXISTS operativos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                semana TEXT, fecha TEXT, tipo TEXT, nombre TEXT,
                direccion TEXT, municipio TEXT, provincia TEXT,
                zona_operativo TEXT, brigadas_requeridas INTEGER DEFAULT 1,
                fuente TEXT DEFAULT 'denuncia', no_oficio TEXT,
                estado TEXT DEFAULT 'pendiente',
                ejecutado INTEGER DEFAULT -1,
                resultado TEXT DEFAULT '',
                observaciones TEXT DEFAULT '',
                decomiso INTEGER DEFAULT 0,
                decomiso_detalle TEXT DEFAULT '',
                brigadas_json TEXT DEFAULT '[]',
                vehiculos_json TEXT DEFAULT '[]',
                seed TEXT, created_at TEXT, confirmed_at TEXT
            );
            CREATE TABLE IF NOT EXISTS denuncias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                no_oficio TEXT, fecha_entrada TEXT, tipo TEXT, nombre TEXT,
                sector TEXT, municipio TEXT, provincia TEXT, zona_inferida TEXT,
                estado TEXT DEFAULT 'pendiente', cargado_semana TEXT
            );
            CREATE TABLE IF NOT EXISTS semanas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                semana TEXT UNIQUE,
                vehiculos_disponibles INTEGER DEFAULT 6,
                militares_disponibles INTEGER DEFAULT 6,
                notas TEXT, estado TEXT DEFAULT 'borrador', created_at TEXT
            );
            CREATE TABLE IF NOT EXISTS sorteo_audit (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                seed TEXT UNIQUE,
                operativo_id INTEGER,
                fecha_sorteo TEXT,
                audit_json TEXT
            );
            CREATE TABLE IF NOT EXISTS uploads_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT, semana TEXT, archivo TEXT, pendientes_cargadas INTEGER
            );
            CREATE TABLE IF NOT EXISTS denuncias_manual (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                no_oficio TEXT,
                fecha_entrada TEXT,
                tipo TEXT,
                nombre TEXT,
                sector TEXT,
                municipio TEXT,
                provincia TEXT,
                direccion TEXT,
                zona_inferida TEXT,
                estado TEXT DEFAULT 'pendiente',
                ingresado_por TEXT,
                usuario_id INTEGER,
                hallazgos TEXT DEFAULT '',
                resolucion TEXT DEFAULT '',
                created_at TEXT
            );
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                nombre_completo TEXT,
                rol TEXT NOT NULL DEFAULT 'coordinador',
                activo INTEGER DEFAULT 1,
                created_at TEXT
            );
        ''')
        # Migration: create sorteo_audit if not exists
        try:
            db.execute('''CREATE TABLE IF NOT EXISTS sorteo_audit (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                seed TEXT UNIQUE,
                operativo_id INTEGER,
                fecha_sorteo TEXT,
                audit_json TEXT
            )''')
        except: pass
        # Migration: new operativos columns
        try:
            db.execute("ALTER TABLE operativos ADD COLUMN denuncia_manual_id INTEGER")
        except: pass
        try:
            db.execute("ALTER TABLE operativos ADD COLUMN via_comunicacion TEXT DEFAULT ''")
        except: pass
        try:
            db.execute("ALTER TABLE operativos ADD COLUMN observacion_orden TEXT DEFAULT ''")
        except: pass
        try:
            db.execute("ALTER TABLE operativos ADD COLUMN resultado_final TEXT DEFAULT ''")
        except: pass
        try:
            db.execute("ALTER TABLE operativos ADD COLUMN bloqueado INTEGER DEFAULT 0")
        except: pass
        try:
            db.execute("ALTER TABLE operativos ADD COLUMN evidencia_path TEXT DEFAULT ''")
        except: pass
        try:
            db.execute("ALTER TABLE operativos ADD COLUMN resultado_evidencia TEXT DEFAULT ''")
        except: pass
        try:
            db.execute("ALTER TABLE personal_state ADD COLUMN motivo_evidencia TEXT DEFAULT ''")
        except: pass
        # Migration: historial_estados table
        try:
            db.execute('''CREATE TABLE IF NOT EXISTS historial_estados (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tabla TEXT NOT NULL,
                registro_id INTEGER NOT NULL,
                fecha TEXT NOT NULL,
                usuario TEXT NOT NULL,
                rol TEXT NOT NULL,
                estado_anterior TEXT,
                estado_nuevo TEXT NOT NULL,
                nota TEXT DEFAULT ''
            )''')
        except: pass
        # Migration: denuncias_manual table
        try:
            db.execute('''CREATE TABLE IF NOT EXISTS denuncias_manual (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                no_oficio TEXT,
                fecha_entrada TEXT,
                tipo TEXT,
                nombre TEXT,
                sector TEXT,
                municipio TEXT,
                provincia TEXT,
                direccion TEXT,
                zona_inferida TEXT,
                estado TEXT DEFAULT 'pendiente',
                ingresado_por TEXT,
                usuario_id INTEGER,
                hallazgos TEXT DEFAULT '',
                resolucion TEXT DEFAULT '',
                created_at TEXT
            )''')
        except: pass
        # Migration: hallazgos/resolucion/estado on denuncias (excel table)
        try:
            db.execute("ALTER TABLE denuncias ADD COLUMN estado TEXT DEFAULT 'pendiente'")
        except: pass
        try:
            db.execute("ALTER TABLE denuncias ADD COLUMN hallazgos TEXT DEFAULT ''")
        except: pass
        try:
            db.execute("ALTER TABLE denuncias ADD COLUMN resolucion TEXT DEFAULT ''")
        except: pass
        # Migration: create uploads dir
        import os as _os
        _os.makedirs('uploads/evidencias', exist_ok=True)
        _os.makedirs('uploads/operativos', exist_ok=True)
        # Seed: create default admin if no users exist
        try:
            u = db.fetchone('SELECT COUNT(*) as c FROM usuarios')
            if not u or u['c'] == 0:
                pw = bcrypt.hashpw(b'admin123', bcrypt.gensalt()).decode()
                db.execute(
                    "INSERT INTO usuarios (username,password_hash,nombre_completo,rol,created_at) VALUES (?,?,?,?,?)",
                    ('admin', pw, 'Administrador del Sistema', 'admin', datetime.now().isoformat())
                )
        except: pass
        # Migration: add estado/hallazgos/resolucion to excel denuncias
        try: db.execute("ALTER TABLE denuncias ADD COLUMN estado TEXT DEFAULT 'pendiente'")
        except: pass
        try: db.execute("ALTER TABLE denuncias ADD COLUMN hallazgos TEXT DEFAULT ''")
        except: pass
        try: db.execute("ALTER TABLE denuncias ADD COLUMN resolucion TEXT DEFAULT ''")
        except: pass
        # Migration: historial_estados - trazabilidad de cambios de estado
        try:
            db.execute('''CREATE TABLE IF NOT EXISTS historial_estados (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tabla TEXT NOT NULL,
                registro_id INTEGER NOT NULL,
                fecha TEXT NOT NULL,
                usuario TEXT NOT NULL,
                rol TEXT NOT NULL,
                estado_anterior TEXT,
                estado_nuevo TEXT NOT NULL,
                nota TEXT DEFAULT ''
            )''')
        except: pass
        # Migration: denuncias_manual
        try:
            db.execute('''CREATE TABLE IF NOT EXISTS denuncias_manual (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                no_oficio TEXT,
                fecha_entrada TEXT,
                tipo TEXT,
                nombre TEXT,
                sector TEXT,
                municipio TEXT,
                provincia TEXT,
                direccion TEXT,
                zona_inferida TEXT,
                estado TEXT DEFAULT 'pendiente',
                ingresado_por TEXT,
                usuario_id INTEGER,
                hallazgos TEXT DEFAULT '',
                resolucion TEXT DEFAULT '',
                created_at TEXT
            )''')
        except: pass
        # Seed personal if empty
        rows = db.fetchall('SELECT COUNT(*) as c FROM personal_state')
        count = rows[0]['c'] if rows else 0
        if count == 0:
            for p in PERSONAL:
                db.execute('INSERT INTO personal_state (persona_id, zona_counts) VALUES (?,?)', (p['id'],'{}'))

try:
    init_db()
except Exception as e:
    import logging
    logging.error(f"init_db failed: {e}")

# ── HELPERS ──────────────────────────────────────────────────
def get_all_states():
    with get_db() as db:
        rows = db.fetchall('SELECT * FROM personal_state')
        return {r['persona_id']: dict(r) for r in rows}

def get_state(pid, all_states=None):
    if all_states:
        return all_states.get(pid, {"persona_id":pid,"carga_total":0,"no_disponible":0,
            "motivo_no_disponible":"","motivo_detalle":"","conflicto":0,
            "ultima_asignacion":None,"zona_counts":"{}"})
    with get_db() as db:
        row = db.fetchone('SELECT * FROM personal_state WHERE persona_id=?', (pid,))
        return dict(row) if row else {"persona_id":pid,"carga_total":0,"no_disponible":0,
            "motivo_no_disponible":"","motivo_detalle":"","conflicto":0,
            "ultima_asignacion":None,"zona_counts":"{}"}

def zona_count(pid, zona, all_states=None):
    st = get_state(pid, all_states)
    return json.loads(st.get('zona_counts') or '{}').get(zona, 0)

def is_elegible(p, zona, fecha, skip_cooldown=False, all_states=None):
    st = get_state(p['id'], all_states)
    LOCAL = ["Distrito Nacional","SD Este","SD Norte","SD Oeste","San Cristóbal","San Cristóbal / Haina"]
    if zona in LOCAL and p['zona'] == zona: return False
    if st['no_disponible']: return False
    if st['conflicto']: return False
    if not skip_cooldown and st['ultima_asignacion']:
        try:
            ref = datetime.fromisoformat(fecha) if fecha else datetime.now()
            d = (ref - datetime.fromisoformat(st['ultima_asignacion'])).days
            if d < COOLDOWN_DAYS: return False
        except: pass
    return True

def select_fair(pool, n, zona, all_states=None):
    if not pool: return []
    scored = sorted(
        [(zona_count(p['id'],zona,all_states), get_state(p['id'],all_states)['carga_total'], p)
         for p in pool],
        key=lambda x:(x[0],x[1])
    )
    mn,mx = scored[0][0],scored[-1][0]; rng=mx-mn
    t = lambda zc: 1 if rng==0 else (1 if zc<=mn+rng/3 else (2 if zc<=mn+2*rng/3 else 3))
    t1=[x[2] for x in scored if t(x[0])==1]; random.shuffle(t1)
    t2=[x[2] for x in scored if t(x[0])==2]; random.shuffle(t2)
    t3=[x[2] for x in scored if t(x[0])==3]; random.shuffle(t3)
    return (t1+t2+t3)[:n]

def registrar_cambio_estado(tabla, registro_id, estado_anterior, estado_nuevo, nota=''):
    """Registra cualquier cambio de estado. Usa autocommit para garantizar que siempre se guarde."""
    import logging
    try:
        usuario = current_user.username if current_user.is_authenticated else 'sistema'
        rol     = current_user.rol      if current_user.is_authenticated else 'sistema'
        # Use get_db_schema (autocommit=True) so this always commits independently
        db = get_db_schema()
        db.execute('''INSERT INTO historial_estados
            (tabla,registro_id,fecha,usuario,rol,estado_anterior,estado_nuevo,nota)
            VALUES (?,?,?,?,?,?,?,?)''',
            (tabla, registro_id, datetime.now().isoformat(),
             usuario, rol, estado_anterior, estado_nuevo, nota))
        db.close()
        logging.info(f"historial_estados: {tabla} #{registro_id} {estado_anterior}→{estado_nuevo}")
    except Exception as e:
        logging.error(f"registrar_cambio_estado FAILED: {e}")

def inferir_zona(prov, mun):
    p = str(prov).upper() if prov and str(prov).lower() not in ['nan','none',''] else ''
    m = str(mun).upper() if mun and str(mun).lower() not in ['nan','none',''] else ''
    if not p: return 'Sin especificar'
    if 'DAJABON' in p or 'DAJABÓN' in p: return 'Dajabón'
    if 'ESPAILLAT' in p: return 'Espaillat'
    if 'SAN JUAN' in p: return 'San Juan'
    if 'SAN PEDRO' in p: return 'San Pedro de Macorís'
    if 'SAN CRISTOBAL' in p or 'SAN CRISTÓBAL' in p:
        return 'San Cristóbal / Haina' if 'HAINA' in m or 'BAJOS' in m else 'San Cristóbal'
    if 'SANTO DOMINGO' in p:
        if 'NORTE' in m: return 'SD Norte'
        if 'ESTE' in m: return 'SD Este'
        if 'OESTE' in m: return 'SD Oeste'
        return 'Distrito Nacional'
    if 'DISTRITO NACIONAL' in p: return 'Distrito Nacional'
    for k,v in [('SANTIAGO','Santiago'),('LA VEGA','La Vega'),('PUERTO PLATA','Puerto Plata'),
                ('LA ROMANA','La Romana'),('BARAHONA','Barahona'),('PERAVIA','Baní / Peravia'),
                ('MONTE PLATA','Monte Plata'),('DUARTE','Duarte'),('ALTAGRACIA','La Altagracia')]:
        if k in p: return v
    return prov.strip().title()

def dias_pendiente(fecha_str):
    try: return (datetime.now() - datetime.fromisoformat(str(fecha_str)[:10])).days
    except: return 0

def row_to_dict(row):
    if row is None: return None
    return dict(row)

# ── ROUTES ────────────────────────────────────────────────────
# ── AUTH ROUTES ──────────────────────────────────────────────
@app.route('/login', methods=['GET','POST'])
def login_view():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','').encode()
        with get_db() as db:
            row = db.fetchone('SELECT * FROM usuarios WHERE username=? AND activo=1', (username,))
        if row and bcrypt.checkpw(password, row['password_hash'].encode()):
            user = User(row['id'], row['username'], row['rol'])
            login_user(user, remember=True)
            if row['rol'] == 'denuncias':
                return redirect(url_for('mis_denuncias_view'))
            return redirect(request.args.get('next') or url_for('index'))
        error = 'Usuario o contrasena incorrectos.'
    return render_template('login.html', error=error)

@app.route('/logout')
@login_required
def logout_view():
    logout_user()
    return redirect(url_for('login_view'))

@app.route('/usuarios')
@rol_required('admin')
def usuarios_view():
    with get_db() as db:
        users = [dict(r) for r in db.fetchall('SELECT * FROM usuarios ORDER BY rol,username')]
    return render_template('usuarios.html', users=users)

@app.route('/usuarios/crear', methods=['POST'])
@rol_required('admin')
def crear_usuario():
    d = request.json
    username = d.get('username','').strip()
    password = d.get('password','').strip()
    nombre   = d.get('nombre','').strip()
    rol      = d.get('rol','coordinador')
    if not username or not password:
        return jsonify(ok=False, error='Usuario y contrasena requeridos'), 400
    pw = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    try:
        with get_db() as db:
            db.execute(
                "INSERT INTO usuarios (username,password_hash,nombre_completo,rol,activo,created_at) VALUES (?,?,?,?,1,?)",
                (username, pw, nombre, rol, datetime.now().isoformat())
            )
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, error='El nombre de usuario ya existe'), 400

@app.route('/usuarios/cambiar_password', methods=['POST'])
@login_required
def cambiar_password():
    d = request.json
    # Admin can change anyone's, others only their own
    target_id = d.get('id', current_user.id)
    if current_user.rol != 'admin' and target_id != current_user.id:
        return jsonify(ok=False, error='Sin permiso'), 403
    pw = bcrypt.hashpw(d['password'].encode(), bcrypt.gensalt()).decode()
    with get_db() as db:
        db.execute('UPDATE usuarios SET password_hash=? WHERE id=?', (pw, target_id))
    return jsonify(ok=True)

@app.route('/usuarios/toggle', methods=['POST'])
@rol_required('admin')
def toggle_usuario():
    d = request.json
    if d['id'] == current_user.id:
        return jsonify(ok=False, error='No puedes desactivarte a ti mismo'), 400
    with get_db() as db:
        db.execute('UPDATE usuarios SET activo=? WHERE id=?', (d['activo'], d['id']))
    return jsonify(ok=True)

@app.route('/usuarios/eliminar/<int:uid>', methods=['DELETE'])
@rol_required('admin')
def eliminar_usuario(uid):
    if uid == current_user.id:
        return jsonify(ok=False, error='No puedes eliminarte a ti mismo'), 400
    with get_db() as db:
        db.execute('DELETE FROM usuarios WHERE id=?', (uid,))
    return jsonify(ok=True)

@app.route('/')
@login_required
def index(): return render_template('index.html')

@app.route('/api/stats')
def api_stats():
    semana = datetime.now().strftime('%Y-W%V')
    with get_db() as db:
        pend = db.fetchone("SELECT COUNT(*) as c FROM denuncias WHERE estado='pendiente'")['c']
        ops  = db.fetchone("SELECT COUNT(*) as c FROM operativos WHERE semana=?", (semana,))['c']
        brig = db.fetchone("SELECT COALESCE(SUM(brigadas_requeridas),0) as c FROM operativos WHERE semana=?", (semana,))['c']
        hist = db.fetchone("SELECT COUNT(*) as c FROM operativos WHERE estado='asignado'")['c']
        sem  = db.fetchone("SELECT * FROM semanas WHERE semana=?", (semana,))
    return jsonify(pendientes=pend, operativos=ops, brigadas=int(brig), historico=hist,
                   vehiculos=sem['vehiculos_disponibles'] if sem else 6,
                   militares=sem['militares_disponibles'] if sem else 6)

@app.route('/personal')
@rol_required('admin')
def personal_view():
    all_states = get_all_states()
    personal = [{**p, **all_states.get(p['id'], {})} for p in PERSONAL]
    with get_db() as db:
        logs = [dict(r) for r in db.fetchall(
            "SELECT * FROM personal_disponibilidad_log ORDER BY fecha_registro DESC LIMIT 50")]
    return render_template('personal.html', personal=personal, cooldown=COOLDOWN_DAYS, logs=logs)

@app.route('/personal/update', methods=['POST'])
@rol_required('admin')
def personal_update():
    d = request.json; pid = d['id']
    with get_db() as db:
        st = db.fetchone('SELECT * FROM personal_state WHERE persona_id=?', (pid,))
        if not st: return jsonify(ok=True)
        nd     = d.get('no_disponible', st['no_disponible'])
        motivo = d.get('motivo_no_disponible', st['motivo_no_disponible'] or '')
        detalle= d.get('motivo_detalle', st['motivo_detalle'] or '')
        ci     = d.get('conflicto', st['conflicto'])
        db.execute('''UPDATE personal_state SET no_disponible=?,motivo_no_disponible=?,
                      motivo_detalle=?,conflicto=? WHERE persona_id=?''',
                   (nd, motivo, detalle, ci, pid))
        if nd != st['no_disponible']:
            db.execute('''INSERT INTO personal_disponibilidad_log
                (persona_id,fecha_registro,motivo,detalle,disponible,semana)
                VALUES (?,?,?,?,?,?)''',
                (pid, datetime.now().isoformat(), motivo, detalle, 1-nd,
                 datetime.now().strftime('%Y-W%V')))
    return jsonify(ok=True)

@app.route('/personal/reset_carga', methods=['POST'])
@rol_required('admin')
def reset_carga():
    with get_db() as db:
        db.execute("UPDATE personal_state SET carga_total=0, zona_counts='{}'")
    return jsonify(ok=True)

@app.route('/denuncias')
@login_required
def denuncias_view():
    with get_db() as db:
        excel_rows = [dict(r) for r in db.fetchall(
            "SELECT * FROM denuncias ORDER BY provincia,municipio")]
        manual_rows = [dict(r) for r in db.fetchall(
            "SELECT * FROM denuncias_manual ORDER BY created_at DESC")]
        logs = [dict(r) for r in db.fetchall(
            "SELECT * FROM uploads_log ORDER BY id DESC LIMIT 10")]
    for r in excel_rows:
        r['zona_inferida'] = inferir_zona(r['provincia'], r['municipio'])
        r['dias'] = dias_pendiente(r['fecha_entrada'])
        r['fuente'] = 'excel'
    for r in manual_rows:
        r['zona_inferida'] = r.get('zona_inferida') or inferir_zona(r.get('provincia',''), r.get('municipio',''))
        r['dias'] = dias_pendiente(r.get('fecha_entrada',''))
        r['fuente'] = 'manual'
    rows = manual_rows + excel_rows
    pendientes = sum(1 for r in rows if r.get('estado') == 'pendiente')
    return render_template('denuncias.html', denuncias=rows, logs=logs, pendientes=pendientes)

@app.route('/denuncias/upload', methods=['POST'])
@rol_required('admin')
def denuncias_upload():
    if 'file' not in request.files: return jsonify(ok=False, error='No file'), 400
    f = request.files['file']
    fname = secure_filename(f.filename)
    path = os.path.join(UPLOAD_FOLDER, fname); f.save(path)
    try:
        df = pd.read_excel(path); df.columns=[c.strip() for c in df.columns]
        rc = next(c for c in df.columns if 'RESOLUCION' in c.upper())
        tc = next(c for c in df.columns if 'TIPO' in c.upper())
        nc = next(c for c in df.columns if 'NOMBRE' in c.upper())
        sc = next(c for c in df.columns if 'SECTOR' in c.upper())
        mc = next(c for c in df.columns if 'MUNICIPIO' in c.upper())
        pc = next(c for c in df.columns if 'PROVINCIA' in c.upper())
        oc = next(c for c in df.columns if 'OFICIO MH' in c.upper())
        fc = next(c for c in df.columns if 'FECHA DE ENTRADA' in c.upper())
        pend = df[df[rc].astype(str).str.upper().str.contains('PENDIENTE', na=False)]
        semana = datetime.now().strftime('%Y-W%V')
        with get_db() as db:
            db.execute("DELETE FROM denuncias WHERE estado='pendiente'")
            for _, row in pend.iterrows():
                pv = str(row[pc]).strip() if pd.notna(row[pc]) else ''
                mu = str(row[mc]).strip() if pd.notna(row[mc]) else ''
                db.execute('''INSERT INTO denuncias
                    (no_oficio,fecha_entrada,tipo,nombre,sector,municipio,provincia,zona_inferida,estado,cargado_semana)
                    VALUES (?,?,?,?,?,?,?,?,?,?)''',
                    (str(row[oc]), str(row[fc])[:10], str(row[tc]), str(row[nc]),
                     str(row[sc]) if pd.notna(row[sc]) else '', mu, pv,
                     inferir_zona(pv, mu), 'pendiente', semana))
            db.execute('INSERT INTO uploads_log (fecha,semana,archivo,pendientes_cargadas) VALUES (?,?,?,?)',
                       (datetime.now().strftime('%Y-%m-%d %H:%M'), semana, fname, len(pend)))
        return jsonify(ok=True, inserted=len(pend))
    except Exception as e: return jsonify(ok=False, error=str(e)), 500

@app.route('/planificacion')
@rol_required('admin')
def planificacion_view():
    semana = datetime.now().strftime('%Y-W%V')
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    week_days = [(monday + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(5)]
    with get_db() as db:
        # Excel-loaded denuncias
        dp_excel = [dict(r) for r in db.fetchall(
            "SELECT *, 'excel' as fuente FROM denuncias WHERE estado='pendiente' ORDER BY provincia,municipio")]
        # Manually entered denuncias (any non-closed state)
        dp_manual = [dict(r) for r in db.fetchall(
            """SELECT *, 'manual' as fuente FROM denuncias_manual
               WHERE estado NOT IN ('cerrada','ejecutada','con_decomiso')
               ORDER BY created_at DESC""")]
        sr  = db.fetchone("SELECT * FROM semanas WHERE semana=?", (semana,))
        ops = [dict(r) for r in db.fetchall(
            "SELECT * FROM operativos WHERE semana=? ORDER BY fecha,id", (semana,))]
    # Combine both sources
    dp = dp_excel + dp_manual
    for d in dp:
        d['zona_inferida'] = inferir_zona(d.get('provincia',''), d.get('municipio',''))
    grupos = {}
    for d in dp:
        z = d['zona_inferida']
        grupos.setdefault(z, []).append(d)
    zonas_agregadas = list({o['zona_operativo'] for o in ops if o['zona_operativo']})
    return render_template('planificacion.html', grupos=grupos, semana=semana,
                           semana_row=row_to_dict(sr), operativos=ops,
                           week_days=week_days, zonas_agregadas=zonas_agregadas)

@app.route('/planificacion/guardar_semana', methods=['POST'])
@rol_required('admin')
def guardar_semana():
    d = request.json
    with get_db() as db:
        db.execute('''INSERT INTO semanas (semana,vehiculos_disponibles,militares_disponibles,created_at)
                      VALUES (?,?,?,?)
                      ON CONFLICT(semana) DO UPDATE SET
                      vehiculos_disponibles=?,militares_disponibles=?''',
                   (d['semana'], int(d.get('vehiculos',6)), int(d.get('militares',6)),
                    datetime.now().isoformat(),
                    int(d.get('vehiculos',6)), int(d.get('militares',6))))
    return jsonify(ok=True)

@app.route('/planificacion/agregar_operativo', methods=['POST'])
@rol_required('admin')
def agregar_operativo():
    d = request.json
    br = 2 if 'DEPORTIVA' in str(d.get('tipo','')).upper() else 1
    den_id = d.get('denuncia_manual_id')
    with get_db() as db:
        db.execute('''INSERT INTO operativos
            (semana,fecha,tipo,nombre,direccion,municipio,provincia,
             zona_operativo,brigadas_requeridas,fuente,no_oficio,created_at,
             denuncia_manual_id,via_comunicacion,observacion_orden)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (d['semana'], d.get('fecha',''), d.get('tipo',''), d.get('nombre',''),
             d.get('direccion',''), d.get('municipio',''), d.get('provincia',''),
             d.get('zona',''), br, d.get('fuente','denuncia'),
             d.get('no_oficio',''), datetime.now().isoformat(), den_id,
             d.get('via_comunicacion',''), d.get('observacion','')))
        if den_id:
            db.execute("UPDATE denuncias_manual SET estado='planificada' WHERE id=?", (den_id,))
            registrar_cambio_estado('denuncias_manual', den_id, 'pendiente', 'planificada',
                                    'Automatico — Incluida en planificacion semanal')
    return jsonify(ok=True)

@app.route('/planificacion/eliminar_operativo/<int:oid>', methods=['DELETE'])
@rol_required('admin')
def eliminar_operativo(oid):
    with get_db() as db:
        db.execute("DELETE FROM operativos WHERE id=?", (oid,))
    return jsonify(ok=True)

@app.route('/asignacion')
@rol_required('admin')
def asignacion_view():
    semana = datetime.now().strftime('%Y-W%V')
    with get_db() as db:
        ops = [dict(r) for r in db.fetchall(
            "SELECT * FROM operativos WHERE semana=? ORDER BY CASE fuente WHEN 'orden_direccion' THEN 0 ELSE 1 END,fecha,id",
            (semana,))]
        sr = db.fetchone("SELECT * FROM semanas WHERE semana=?", (semana,))
    all_states = get_all_states()
    disponibles = sum(1 for p in PERSONAL
                      if not all_states.get(p['id'], {}).get('no_disponible', 0)
                      and not all_states.get(p['id'], {}).get('conflicto', 0))
    tb = sum(o['brigadas_requeridas'] for o in ops)
    v  = sr['vehiculos_disponibles'] if sr else 6
    m  = sr['militares_disponibles'] if sr else 6
    return render_template('asignacion.html', operativos=ops, semana=semana,
                           vehiculos=v, militares=m, total_brigadas=tb,
                           disponibles=disponibles)

@app.route('/asignacion/ejecutar', methods=['POST'])
@rol_required('admin')
def ejecutar_asignacion():
    d = request.json; semana = d['semana']; veh = int(d.get('vehiculos', 6))
    fb = d.get('fecha_base', datetime.now().strftime('%Y-%m-%d'))
    with get_db() as db:
        ops = [dict(r) for r in db.fetchall(
            "SELECT * FROM operativos WHERE semana=? AND estado='pendiente' ORDER BY CASE fuente WHEN 'orden_direccion' THEN 0 ELSE 1 END,fecha,id",
            (semana,))]
    # Trim to vehicle budget
    sel, used = [], 0
    for o in ops:
        if used + o['brigadas_requeridas'] <= veh:
            sel.append(o); used += o['brigadas_requeridas']

    all_states = get_all_states()
    coords = [p for p in PERSONAL if p['cargo']=='COORDINADOR']
    insps  = [p for p in PERSONAL if p['cargo']=='INSPECTOR']
    auxs   = [p for p in PERSONAL if p['cargo']=='AUXILIAR']
    arun = set(); aday = {}
    seed = secrets.token_hex(4).upper(); resultado = []

    for op in sel:
        zona  = op['zona_operativo'] or inferir_zona(op['provincia'], op['municipio'])
        fecha = op['fecha'] or fb
        bops  = []; cooldown_relajado = False

        for b in range(op['brigadas_requeridas']):
            def pool(lst, skip_cd=False):
                return [p for p in lst
                        if p['id'] not in arun
                        and (fecha not in aday or p['id'] not in aday[fecha])
                        and is_elegible(p, zona, fecha, skip_cooldown=skip_cd, all_states=all_states)]

            cp = pool(coords); ip = pool(insps); ap = pool(auxs)
            if not cp or not ip or not ap:
                cp2 = pool(coords, True); ip2 = pool(insps, True); ap2 = pool(auxs, True)
                if not cp and cp2: cp = cp2; cooldown_relajado = True
                if not ip and ip2: ip = ip2; cooldown_relajado = True
                if not ap and ap2: ap = ap2; cooldown_relajado = True

            c = select_fair(cp, 1, zona, all_states)
            i = select_fair(ip, 1, zona, all_states)
            a = select_fair(ap, 1, zona, all_states)
            bops.append({"num": b+1,
                         "coordinador": c[0] if c else None,
                         "inspector":   i[0] if i else None,
                         "auxiliar":    a[0] if a else None,
                         "cooldown_relajado": cooldown_relajado})
            for p in filter(None, [c[0] if c else None, i[0] if i else None, a[0] if a else None]):
                arun.add(p['id']); aday.setdefault(fecha, set()).add(p['id'])

        resultado.append({**op, 'brigadas_asignadas': bops, 'seed': seed,
                          'cooldown_relajado': cooldown_relajado})
    return jsonify(ok=True, resultado=resultado, seed=seed)

@app.route('/asignacion/confirmar', methods=['POST'])
@rol_required('admin')
def confirmar_asignacion():
    now = datetime.now().isoformat()
    resultado = request.json['resultado']
    # Load states ONCE outside the DB transaction
    all_states = get_all_states()

    for op in resultado:
        brigadas = op['brigadas_asignadas']
        seed     = op.get('seed','')
        zona     = op.get('zona_operativo','') or inferir_zona(op.get('provincia',''), op.get('municipio',''))
        fecha    = op.get('fecha', datetime.now().strftime('%Y-%m-%d'))

        # Build audit record using already-loaded states
        asignados = []
        excluidos_detail = []
        LOCAL = ["Distrito Nacional","SD Este","SD Norte","SD Oeste","San Cristóbal","San Cristóbal / Haina"]
        for p in PERSONAL:
            st = get_state(p['id'], all_states)
            excluded = False
            reasons = []
            if zona in LOCAL and p['zona']==zona:
                excluded=True; reasons.append("Zona residencia")
            if st['no_disponible']:
                excluded=True; reasons.append(st.get('motivo_no_disponible') or "No disponible")
            if st['conflicto']:
                excluded=True; reasons.append("Conflicto")
            if st['ultima_asignacion']:
                try:
                    ref = datetime.fromisoformat(fecha)
                    d = (ref - datetime.fromisoformat(st['ultima_asignacion'])).days
                    if d < COOLDOWN_DAYS:
                        excluded=True; reasons.append(f"Cooldown {d}d")
                except: pass
            if excluded:
                excluidos_detail.append({"id":p['id'],"nombre":p['nombre'],"cargo":p['cargo'],"razones":reasons})
        for b in brigadas:
            for role in ['coordinador','inspector','auxiliar']:
                p = b.get(role)
                if p: asignados.append({"id":p['id'],"nombre":p['nombre'],"cargo":p['cargo'],"vehiculo":b['num']})

        audit_data = json.dumps({
            "seed": seed, "operativo_id": op['id'],
            "operativo_nombre": op.get('nombre',''),
            "zona": zona, "fecha": fecha, "timestamp": now,
            "asignados": asignados, "excluidos": excluidos_detail,
            "total_pool": len(PERSONAL),
            "total_elegibles": len(PERSONAL) - len(excluidos_detail),
        })

        with get_db() as db:
            db.execute("""UPDATE operativos SET brigadas_json=?,seed=?,estado='asignado',confirmed_at=?
                          WHERE id=?""", (json.dumps(brigadas), seed, now, op['id']))
            # Use seed+operativo_id as unique key to avoid conflicts
            try:
                db.execute("""INSERT INTO sorteo_audit (seed, operativo_id, fecha_sorteo, audit_json)
                              VALUES (?,?,?,?)""", (seed+'-'+str(op['id']), op['id'], now, audit_data))
            except: pass
            for b in brigadas:
                for role in ['coordinador','inspector','auxiliar']:
                    p = b.get(role)
                    if not p: continue
                    pid = p['id']
                    st  = get_state(pid, all_states)
                    zc  = json.loads(st.get('zona_counts') or '{}')
                    zc[zona] = zc.get(zona, 0) + 1
                    db.execute('''UPDATE personal_state SET carga_total=carga_total+1,
                                  ultima_asignacion=?,zona_counts=? WHERE persona_id=?''',
                               (fecha, json.dumps(zc), pid))
    return jsonify(ok=True)

@app.route('/operativo/resultado', methods=['POST'])
@rol_required('admin')
def guardar_resultado():
    d = request.json
    ejecutado = d.get('ejecutado', -1)
    decomiso  = d.get('decomiso', 0)
    resultado_final = d.get('resultado_final','')  # 'ejecutado','ejecutado_sin_incautacion','no_ejecutado'

    # Determine denuncia estado and whether to block operativo
    if resultado_final == 'ejecutado':
        nuevo_estado_den = 'ejecutada'
        bloquear = 1  # Cannot change after executed with incautacion
    elif resultado_final == 'con_decomiso':
        nuevo_estado_den = 'con_decomiso'
        bloquear = 1
    elif resultado_final == 'ejecutado_sin_incautacion':
        nuevo_estado_den = 'pendiente'  # Stays pending for reassignment
        bloquear = 1  # This specific operativo is locked
    elif resultado_final == 'no_ejecutado':
        nuevo_estado_den = 'pendiente'  # Stays pending
        bloquear = 0
    else:
        nuevo_estado_den = None
        bloquear = 0

    op = None
    den_id = None
    anterior_estado = None

    with get_db() as db:
        # Check if already blocked
        op = db.fetchone('SELECT * FROM operativos WHERE id=?', (d['id'],))
        if op and op.get('bloqueado') == 1:
            return jsonify(ok=False, error='Este operativo ya fue registrado y no puede modificarse.'), 403

        # Update core fields
        db.execute(
            'UPDATE operativos SET ejecutado=?,resultado=?,observaciones=?,decomiso=?,decomiso_detalle=? WHERE id=?',
            (ejecutado, d.get('resultado',''), d.get('observaciones',''),
             decomiso, d.get('decomiso_detalle',''), d['id']))

        # Update new fields separately
        try:
            db.execute('UPDATE operativos SET resultado_final=?,bloqueado=? WHERE id=?',
                       (resultado_final, bloquear, d['id']))
        except: pass

        # Update linked denuncia_manual
        den_id = op.get('denuncia_manual_id') if op else None
        if nuevo_estado_den and den_id:
            try:
                actual = db.fetchone('SELECT estado FROM denuncias_manual WHERE id=?', (den_id,))
                anterior_estado = actual['estado'] if actual else None
                db.execute('UPDATE denuncias_manual SET estado=? WHERE id=?',
                           (nuevo_estado_den, den_id))
            except: pass

    # AFTER commit: register historial (autocommit — always saves)
    if den_id and resultado_final:
        estado_log = resultado_final
        nota_auto = f"Ejecucion Diaria — {estado_log.replace('_',' ').upper()}"
        if d.get('observaciones'):
            nota_auto += f": {d.get('observaciones','')[:100]}"
        registrar_cambio_estado('denuncias_manual', den_id,
            anterior_estado, estado_log, nota_auto)

    return jsonify(ok=True)

@app.route('/denuncias_manual/actualizar', methods=['POST'])
@login_required
def actualizar_denuncia_manual():
    if current_user.rol not in ['admin','operaciones']:
        return jsonify(ok=False, error='Sin permiso'), 403
    d = request.json
    fuente   = d.get('fuente','manual')
    nuevo    = d['estado']
    hallazgos= d.get('hallazgos','')
    resolucion=d.get('resolucion','')
    nota     = d.get('nota','')
    # Block system-only states from being set manually
    if nuevo in ('ejecutada','con_decomiso','cerrada'):
        return jsonify(ok=False, error='Este estado solo lo puede asignar el sistema via Ejecucion Diaria.'), 403
    with get_db() as db:
        if fuente == 'excel':
            actual = db.fetchone('SELECT estado FROM denuncias WHERE id=?', (d['id'],))
            anterior = actual['estado'] if actual else None
            db.execute("UPDATE denuncias SET estado=?,hallazgos=?,resolucion=? WHERE id=?",
                (nuevo, hallazgos, resolucion, d['id']))
            registrar_cambio_estado('denuncias', d['id'], anterior, nuevo,
                nota or hallazgos[:80] if hallazgos else '')
        else:
            actual = db.fetchone('SELECT estado FROM denuncias_manual WHERE id=?', (d['id'],))
            anterior = actual['estado'] if actual else None
            db.execute("UPDATE denuncias_manual SET estado=?,hallazgos=?,resolucion=? WHERE id=?",
                (nuevo, hallazgos, resolucion, d['id']))
            registrar_cambio_estado('denuncias_manual', d['id'], anterior, nuevo,
                nota or hallazgos[:80] if hallazgos else '')
    return jsonify(ok=True)

@app.route('/operativo/vehiculos', methods=['POST'])
@rol_required('admin')
def guardar_vehiculos_operativo():
    d = request.json
    with get_db() as db:
        db.execute("UPDATE operativos SET vehiculos_json=? WHERE id=?",
                   (json.dumps(d['vehiculos']), d['id']))
    return jsonify(ok=True)

@app.route('/plan_semanal')
@login_required
def plan_semanal():
    semana = request.args.get('semana', datetime.now().strftime('%Y-W%V'))
    with get_db() as db:
        ops = [dict(r) for r in db.fetchall(
            "SELECT * FROM operativos WHERE semana=? ORDER BY fecha,id", (semana,))]
        sr  = db.fetchone("SELECT * FROM semanas WHERE semana=?", (semana,))
    for o in ops:
        o['brigadas']      = json.loads(o['brigadas_json'] or '[]')
        o['vehiculos_data']= json.loads(o['vehiculos_json'] or '[]')
    return render_template('plan_semanal.html', operativos=ops, semana=semana,
                           semana_row=row_to_dict(sr))

@app.route('/ejecucion_diaria')
@login_required
def ejecucion_diaria():
    today  = datetime.now().strftime('%Y-%m-%d')
    # Always use today - sorteo only happens day-of
    fecha  = today
    semana = datetime.now().strftime('%Y-W%V')
    with get_db() as db:
        ops     = [dict(r) for r in db.fetchall(
            "SELECT * FROM operativos WHERE fecha=? ORDER BY id", (fecha,))]
        all_ops = [dict(r) for r in db.fetchall(
            "SELECT DISTINCT fecha FROM operativos WHERE semana=? ORDER BY fecha", (semana,))]
        sr = db.fetchone("SELECT * FROM semanas WHERE semana=?", (semana,))
    for o in ops:
        o['brigadas']       = json.loads(o['brigadas_json'] or '[]')
        o['vehiculos_data'] = json.loads(o['vehiculos_json'] or '[]')
    vehiculos_diarios = sr['vehiculos_disponibles'] if sr else 6
    return render_template('ejecucion_diaria.html', operativos=ops, fecha=fecha,
                           fechas_disponibles=[x['fecha'] for x in all_ops],
                           semana=semana, vehiculos_diarios=vehiculos_diarios,
                           today=today)

@app.route('/historial')
@login_required
def historial_view():
    semana = request.args.get('semana', datetime.now().strftime('%Y-W%V'))
    with get_db() as db:
        ops_semana = [dict(r) for r in db.fetchall(
            "SELECT * FROM operativos WHERE semana=? ORDER BY fecha,id", (semana,))]
        ops_all = [dict(r) for r in db.fetchall(
            "SELECT * FROM operativos WHERE ejecutado != -1 OR resultado_final != '' ORDER BY fecha DESC,id DESC")]
        semanas_list = [dict(r) for r in db.fetchall(
            "SELECT DISTINCT semana FROM operativos ORDER BY semana DESC")]
    for o in ops_semana: o['brigadas'] = json.loads(o['brigadas_json'] or '[]')
    for o in ops_all:    o['brigadas'] = json.loads(o['brigadas_json'] or '[]')
    # Weekly stats
    dias = ['Lunes','Martes','Miercoles','Jueves','Viernes']
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    week_dates = [(monday + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(5)]
    daily = []
    for i,(dia,fecha) in enumerate(zip(dias,week_dates)):
        plan    = [o for o in ops_semana if o['fecha']==fecha]
        ejec    = [o for o in plan if o.get('resultado_final') == 'ejecutado' or (o['ejecutado']==1 and o.get('resultado_final') not in ('ejecutado_sin_incautacion','con_decomiso',''))]
        sin_inc = [o for o in plan if o.get('resultado_final') == 'ejecutado_sin_incautacion']
        decomiso= [o for o in plan if o.get('resultado_final') == 'con_decomiso']
        pend    = [o for o in plan if o['ejecutado'] not in (0,1)]
        pct     = round((len(ejec)+len(sin_inc)+len(decomiso))/len(plan)*100) if plan else 0
        daily.append({'dia':dia,'fecha':fecha,'planificadas':len(plan),
                      'ejecutadas':len(ejec),'sin_incautacion':len(sin_inc),
                      'con_decomiso':len(decomiso),'pendientes':len(pend),'pct':pct})
    total_plan    = sum(d['planificadas'] for d in daily)
    total_ejec    = sum(d['ejecutadas'] for d in daily)
    total_sin_inc = sum(d['sin_incautacion'] for d in daily)
    total_pend    = sum(d['pendientes'] for d in daily)
    total_pct     = round((total_ejec+total_sin_inc)/total_plan*100) if total_plan else 0
    return render_template('historial.html',
        operativos=ops_all, ops_semana=ops_semana, semana=semana,
        semanas_list=[s['semana'] for s in semanas_list],
        daily=daily, total_plan=total_plan, total_ejec=total_ejec,
        total_sin_inc=total_sin_inc, total_pend=total_pend, total_pct=total_pct)

@app.route('/reportes')
@login_required
def reportes_view():
    filtro = request.args.get('filtro', 'semana')
    valor  = request.args.get('valor', datetime.now().strftime('%Y-W%V'))
    with get_db() as db:
        if filtro == 'dia':
            ops = [dict(r) for r in db.fetchall(
                "SELECT * FROM operativos WHERE fecha=? AND estado='asignado'", (valor,))]
        elif filtro == 'mes':
            ops = [dict(r) for r in db.fetchall(
                "SELECT * FROM operativos WHERE substr(fecha,1,7)=? AND estado='asignado'", (valor,))]
        else:
            ops = [dict(r) for r in db.fetchall(
                "SELECT * FROM operativos WHERE semana=? AND estado='asignado'", (valor,))]
        denuncias_all = [dict(r) for r in db.fetchall("SELECT * FROM denuncias ORDER BY fecha_entrada")]
    for d in denuncias_all:
        d['dias'] = dias_pendiente(d['fecha_entrada'])
    for o in ops:
        o['brigadas'] = json.loads(o['brigadas_json'] or '[]')
    all_states = get_all_states()
    rendimiento = [{**p, 'st': all_states.get(p['id'], {"carga_total":0,"ultima_asignacion":None})}
                   for p in PERSONAL]
    rendimiento.sort(key=lambda x: -x['st']['carga_total'])
    # Operativos por provincia
    por_prov = {}
    for o in ops:
        pv = o['provincia'] or 'Sin especificar'
        por_prov[pv] = por_prov.get(pv, 0) + 1
    por_prov_sorted = sorted(por_prov.items(), key=lambda x:-x[1])
    max_prov = max(por_prov.values()) if por_prov else 1

    # Denuncias pendientes por zona
    pend_zona = {}
    for d in denuncias_all:
        if d['estado'] == 'pendiente':
            z = d['zona_inferida'] or 'Sin especificar'
            pend_zona[z] = pend_zona.get(z, 0) + 1
    pend_zona_sorted = sorted(pend_zona.items(), key=lambda x:-x[1])
    max_pend = max(pend_zona.values()) if pend_zona else 1

    return render_template('reportes.html',
        ops=ops,
        ejecutados   =[o for o in ops if o.get('ejecutado')==1],
        no_ejecutados=[o for o in ops if o.get('ejecutado')==0],
        con_decomiso =[o for o in ops if o.get('decomiso')==1],
        denuncias=denuncias_all,
        rendimiento=rendimiento,
        por_prov=por_prov_sorted, max_prov=max_prov,
        pend_zona=pend_zona_sorted, max_pend=max_pend,
        filtro=filtro, valor=valor,
        total_pendientes=sum(1 for d in denuncias_all if d['estado']=='pendiente'))

@app.route('/auditoria')
@login_required
def auditoria_view():
    q = request.args.get('q','').strip().upper()
    resultado = None
    if q:
        with get_db() as db:
            row = db.fetchone("SELECT * FROM sorteo_audit WHERE UPPER(seed) LIKE ?", (q+'%',))
            if row:
                resultado = dict(row)
                resultado['data'] = json.loads(resultado['audit_json'])
    return render_template('auditoria.html', q=q, resultado=resultado)

# ── DENUNCIAS DPTO ROUTES ─────────────────────────────────────
@app.route('/mis_denuncias')
@login_required
def mis_denuncias_view():
    """Vista exclusiva del perfil denuncias."""
    if current_user.rol not in ['denuncias']:
        return redirect(url_for('index'))
    with get_db() as db:
        # Manual denuncias ingresadas por este usuario
        manual_rows = [dict(r) for r in db.fetchall(
            "SELECT * FROM denuncias_manual WHERE usuario_id=? ORDER BY created_at DESC",
            (current_user.id,))]
        # Excel denuncias (visible to all denuncias profiles)
        excel_rows = [dict(r) for r in db.fetchall(
            "SELECT * FROM denuncias ORDER BY fecha_entrada DESC")]
    for r in manual_rows:
        r['fuente'] = 'manual'
        r['dias'] = dias_pendiente(r.get('fecha_entrada',''))
    for r in excel_rows:
        r['fuente'] = 'excel'
        r['dias'] = dias_pendiente(r.get('fecha_entrada',''))
        r['zona_inferida'] = inferir_zona(r.get('provincia',''), r.get('municipio',''))
    # Merge: manual first then excel
    rows = manual_rows + excel_rows
    return render_template('mis_denuncias.html', denuncias=rows)

@app.route('/mis_denuncias/ingresar', methods=['POST'])
@login_required
def ingresar_denuncia():
    if current_user.rol not in ['denuncias','admin']:
        return jsonify(ok=False, error='Sin permiso'), 403
    d = request.json
    nombre   = d.get('nombre','').strip()
    provincia= d.get('provincia','').strip()
    if not nombre or not provincia:
        return jsonify(ok=False, error='Nombre y provincia son requeridos'), 400
    municipio = d.get('municipio','').strip()
    with get_db() as db:
        db.execute("""INSERT INTO denuncias_manual
            (no_oficio,fecha_entrada,tipo,nombre,sector,municipio,provincia,
             direccion,zona_inferida,estado,ingresado_por,usuario_id,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (d.get('no_oficio','').strip(),
             datetime.now().strftime('%Y-%m-%d'),
             d.get('tipo','BANCAS DE LOTERIA'), nombre,
             d.get('sector','').strip(), municipio, provincia,
             d.get('direccion','').strip(),
             inferir_zona(provincia, municipio),
             'pendiente', current_user.username, current_user.id,
             datetime.now().isoformat()))
    return jsonify(ok=True)

@app.route('/mis_denuncias/exportar')
@login_required
def exportar_mis_denuncias():
    if current_user.rol not in ['denuncias','admin']:
        return jsonify(ok=False), 403
    import io
    from flask import send_file
    with get_db() as db:
        if current_user.rol == 'denuncias':
            manual = [dict(r) for r in db.fetchall(
                "SELECT * FROM denuncias_manual WHERE usuario_id=? ORDER BY created_at DESC",
                (current_user.id,))]
            excel  = [dict(r) for r in db.fetchall(
                "SELECT * FROM denuncias ORDER BY fecha_entrada DESC")]
            rows = manual + excel
        else:
            rows = [dict(r) for r in db.fetchall(
                "SELECT * FROM denuncias_manual ORDER BY created_at DESC")]
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = 'Denuncias'
        hfill = PatternFill('solid', fgColor='0D2257')
        hfont = Font(color='FFFFFF', bold=True, size=10)
        cols = ['#','No. Oficio','Fecha','Tipo','Nombre','Sector',
                'Municipio','Provincia','Direccion','Zona','Estado','Ingresado por']
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal='center')
        for ri, r in enumerate(rows, 2):
            vals = [ri-1, r.get('no_oficio',''), r.get('fecha_entrada',''),
                    r.get('tipo',''), r.get('nombre',''), r.get('sector',''),
                    r.get('municipio',''), r.get('provincia',''),
                    r.get('direccion',''), r.get('zona_inferida',''),
                    r.get('estado',''), r.get('ingresado_por','')]
            for ci, v in enumerate(vals, 1):
                ws.cell(row=ri, column=ci, value=v)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max((len(str(c.value or '')) for c in col), default=8) + 4, 40)
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"Denuncias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return send_file(buf, download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True)
    except Exception as e:
        return jsonify(ok=False, error=str(e)), 500

@app.route('/operaciones/denuncias')
@login_required
def operaciones_denuncias_view():
    """Vista del admin/operaciones para gestionar denuncias del dpto."""
    if current_user.rol not in ['admin','operaciones']:
        return redirect(url_for('index'))
    with get_db() as db:
        rows = [dict(r) for r in db.fetchall(
            "SELECT * FROM denuncias_manual ORDER BY CASE estado WHEN 'pendiente' THEN 0 WHEN 'planificada' THEN 1 WHEN 'en_ejecucion' THEN 2 ELSE 3 END, created_at DESC")]
    for r in rows:
        r['dias'] = dias_pendiente(r.get('fecha_entrada',''))
    pendientes = sum(1 for r in rows if r['estado'] == 'pendiente')
    return render_template('operaciones_denuncias.html', denuncias=rows, pendientes=pendientes)

@app.route('/operaciones/denuncias/actualizar', methods=['POST'])
@login_required
def actualizar_estado_denuncia():
    if current_user.rol not in ['admin','operaciones']:
        return jsonify(ok=False, error='Sin permiso'), 403
    d = request.json
    did = d['id']; estado = d['estado']
    hallazgos  = d.get('hallazgos','').strip()
    resolucion = d.get('resolucion','').strip()
    with get_db() as db:
        db.execute(
            "UPDATE denuncias_manual SET estado=?,hallazgos=?,resolucion=? WHERE id=?",
            (estado, hallazgos, resolucion, did))
    return jsonify(ok=True)

@app.route('/denuncias/historial/<fuente>/<int:did>')
@login_required
def historial_denuncia(fuente, did):
    tabla = 'denuncias' if fuente == 'excel' else 'denuncias_manual'
    with get_db() as db:
        # Estado change log
        logs = [dict(r) for r in db.fetchall(
            "SELECT * FROM historial_estados WHERE tabla=? AND registro_id=? ORDER BY fecha",
            (tabla, did))]
        # Linked operativos (only for manual denuncias)
        operativos = []
        if fuente == 'manual':
            ops = [dict(r) for r in db.fetchall(
                """SELECT o.id, o.fecha, o.zona_operativo, o.resultado_final,
                          o.observaciones, o.ejecutado, o.decomiso, o.brigadas_json
                   FROM operativos o
                   WHERE o.denuncia_manual_id=?
                   ORDER BY o.fecha ASC""", (did,))]
            for op in ops:
                try:
                    op['brigadas'] = json.loads(op.get('brigadas_json') or '[]')
                except:
                    op['brigadas'] = []
            operativos = ops
    return jsonify(ok=True, historial=logs, operativos=operativos)

@app.route('/denuncias/visitas/<int:did>')
@login_required
def visitas_denuncia(did):
    """Historial de visitas (operativos) para una denuncia manual."""
    with get_db() as db:
        den = db.fetchone('SELECT * FROM denuncias_manual WHERE id=?', (did,))
        if not den:
            return jsonify(ok=False, error='No encontrada'), 404
        ops = [dict(r) for r in db.fetchall(
            """SELECT o.*, s.seed FROM operativos o
               LEFT JOIN sorteo_audit s ON s.operativo_id=o.id
               WHERE o.denuncia_manual_id=?
               ORDER BY o.fecha DESC""", (did,))]
        # For each operativo, get assigned personal
        for op in ops:
            brigadas = []
            try:
                brigadas = json.loads(op.get('brigadas_json') or '[]')
            except: pass
            op['brigadas'] = brigadas
            op['sin_incautacion'] = op.get('resultado_final') == 'ejecutado_sin_incautacion'
    visitas_sin_incautacion = sum(1 for o in ops if o.get('resultado_final') == 'ejecutado_sin_incautacion')
    return jsonify(ok=True, denuncia=dict(den), operativos=ops,
                   visitas_sin_incautacion=visitas_sin_incautacion)

@app.route('/planificacion/agregar_orden_con_evidencia', methods=['POST'])
@rol_required('admin')
def agregar_orden_con_evidencia():
    import os, uuid
    tipo    = request.form.get('tipo','')
    nombre  = request.form.get('nombre','').strip()
    semana  = request.form.get('semana','')
    fecha   = request.form.get('fecha','')
    prov    = request.form.get('provincia','').strip()
    municipio=request.form.get('municipio','').strip()
    direccion=request.form.get('direccion','').strip()
    zona    = request.form.get('zona','')
    via     = request.form.get('via_comunicacion','')
    obs     = request.form.get('observacion','').strip()
    br      = 2 if 'DEPORTIVA' in tipo.upper() else 1
    # Save evidence file
    ev_file = request.files.get('evidencia')
    ev_path = ''
    if ev_file and ev_file.filename:
        ext = ev_file.filename.rsplit('.',1)[-1].lower() if '.' in ev_file.filename else 'jpg'
        fname = f"{uuid.uuid4().hex}.{ext}"
        os.makedirs('uploads/evidencias', exist_ok=True)
        ev_file.save(f"uploads/evidencias/{fname}")
        ev_path = f"evidencias/{fname}"
    with get_db() as db:
        db.execute('''INSERT INTO operativos
            (semana,fecha,tipo,nombre,direccion,municipio,provincia,
             zona_operativo,brigadas_requeridas,fuente,no_oficio,created_at,
             via_comunicacion,observacion_orden,evidencia_path)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (semana,fecha,tipo,nombre,direccion,municipio,prov,
             zona,br,'orden_direccion','',datetime.now().isoformat(),
             via,obs,ev_path))
    return jsonify(ok=True)

@app.route('/evidencia/<path:filename>')
@login_required
def ver_evidencia(filename):
    from flask import send_from_directory
    return send_from_directory('uploads', filename)

@app.route('/buscar')
@login_required
def buscar_view():
    q = request.args.get('q','').strip()
    resultados = []
    if q:
        with get_db() as db:
            ops = [dict(r) for r in db.fetchall(
                """SELECT * FROM operativos
                   WHERE LOWER(nombre) LIKE LOWER(?) OR LOWER(no_oficio) LIKE LOWER(?)
                      OR LOWER(municipio) LIKE LOWER(?) OR LOWER(provincia) LIKE LOWER(?)
                      OR LOWER(zona_operativo) LIKE LOWER(?)
                   ORDER BY fecha DESC LIMIT 50""",
                (f'%{q}%',f'%{q}%',f'%{q}%',f'%{q}%',f'%{q}%'))]
            for op in ops:
                try: op['brigadas'] = json.loads(op.get('brigadas_json') or '[]')
                except: op['brigadas'] = []
            resultados = ops
    return render_template('buscar.html', q=q, resultados=resultados)

@app.route('/operativo/resultado_con_evidencia', methods=['POST'])
@rol_required('admin')
def guardar_resultado_con_evidencia():
    import uuid, os
    op_id = int(request.form.get('id',0))
    resultado_final = request.form.get('resultado_final','')
    observaciones   = request.form.get('observaciones','')
    decomiso        = int(request.form.get('decomiso',0))
    decomiso_detalle= request.form.get('decomiso_detalle','')
    ejecutado = 1 if resultado_final in ('ejecutado','ejecutado_sin_incautacion','con_decomiso') else 0
    bloquear  = 1 if resultado_final in ('ejecutado','con_decomiso','ejecutado_sin_incautacion') else 0
    nuevo_estado_den = 'ejecutada' if resultado_final=='ejecutado' else                        'con_decomiso' if resultado_final=='con_decomiso' else                        'pendiente' if resultado_final=='ejecutado_sin_incautacion' else 'pendiente'
    # Save evidence file
    ev_file = request.files.get('resultado_evidencia')
    ev_path = ''
    if ev_file and ev_file.filename:
        ext = ev_file.filename.rsplit('.',1)[-1].lower() if '.' in ev_file.filename else 'jpg'
        fname = f"op_{op_id}_{uuid.uuid4().hex[:8]}.{ext}"
        os.makedirs('uploads/operativos', exist_ok=True)
        ev_file.save(f"uploads/operativos/{fname}")
        ev_path = f"operativos/{fname}"
    op = None
    den_id = None
    anterior_estado = None
    with get_db() as db:
        op = db.fetchone('SELECT * FROM operativos WHERE id=?', (op_id,))
        if op and op.get('bloqueado') == 1:
            return jsonify(ok=False, error='Este operativo ya fue registrado y no puede modificarse.'), 403
        db.execute(
            'UPDATE operativos SET ejecutado=?,resultado=?,observaciones=?,decomiso=?,decomiso_detalle=? WHERE id=?',
            (ejecutado, resultado_final, observaciones, decomiso, decomiso_detalle, op_id))
        try:
            db.execute('UPDATE operativos SET resultado_final=?,bloqueado=?,resultado_evidencia=? WHERE id=?',
                       (resultado_final, bloquear, ev_path, op_id))
        except: pass
        den_id = op.get('denuncia_manual_id') if op else None
        if nuevo_estado_den and den_id:
            try:
                actual = db.fetchone('SELECT estado FROM denuncias_manual WHERE id=?', (den_id,))
                anterior_estado = actual['estado'] if actual else None
                db.execute('UPDATE denuncias_manual SET estado=? WHERE id=?', (nuevo_estado_den, den_id))
            except: pass
    if den_id and resultado_final:
        nota = f"Ejecucion Diaria — {resultado_final.replace('_',' ').upper()}: {observaciones[:100]}"
        registrar_cambio_estado('denuncias_manual', den_id, anterior_estado, resultado_final, nota)
    return jsonify(ok=True, evidencia=ev_path)

@app.route('/personal/update_con_evidencia', methods=['POST'])
@rol_required('admin')
def personal_update_con_evidencia():
    import uuid, os
    pid = int(request.form.get('id',0))
    disponible = int(request.form.get('disponible',1))
    motivo = request.form.get('motivo','').strip()
    detalle = request.form.get('detalle','').strip()
    if not disponible and not motivo:
        return jsonify(ok=False, error='El motivo es obligatorio para marcar como no disponible'), 400
    # Save evidence
    ev_file = request.files.get('evidencia')
    ev_path = ''
    if ev_file and ev_file.filename:
        ext = ev_file.filename.rsplit('.',1)[-1].lower() if '.' in ev_file.filename else 'pdf'
        fname = f"personal_{pid}_{uuid.uuid4().hex[:8]}.{ext}"
        os.makedirs('uploads/evidencias', exist_ok=True)
        ev_file.save(f"uploads/evidencias/{fname}")
        ev_path = f"evidencias/{fname}"
    with get_db() as db:
        try:
            db.execute('''UPDATE personal_state
                SET disponible=?,motivo_no_disponible=?,motivo_detalle=?,motivo_evidencia=?
                WHERE personal_id=?''',
                (disponible, motivo if not disponible else '', detalle, ev_path if ev_path else '', pid))
        except:
            db.execute('''UPDATE personal_state
                SET disponible=?,motivo_no_disponible=?,motivo_detalle=?
                WHERE personal_id=?''',
                (disponible, motivo if not disponible else '', detalle, pid))
        # Log
        nombre = db.fetchone('SELECT nombre FROM personal_state WHERE personal_id=?',(pid,))
        nom = nombre['nombre'] if nombre else str(pid)
    estado_log = 'no_disponible' if not disponible else 'disponible'
    registrar_cambio_estado('personal_state', pid, None, estado_log,
        f"{nom} — Motivo: {motivo} — {detalle}")
    return jsonify(ok=True)

@app.route('/mapa')
@login_required
def mapa_view():
    with get_db() as db:
        denuncias = [dict(r) for r in db.fetchall("SELECT * FROM denuncias WHERE estado='pendiente'")]
    puntos = []
    for d in denuncias:
        zona   = d['zona_inferida'] or inferir_zona(d['provincia'], d['municipio'])
        coords = PROV_COORDS.get(zona) or PROV_COORDS.get(d['provincia'])
        if coords:
            puntos.append({'lat': coords[0]+random.uniform(-0.05,0.05),
                           'lng': coords[1]+random.uniform(-0.05,0.05),
                           'nombre': d['nombre'], 'tipo': d['tipo'],
                           'municipio': d['municipio'], 'provincia': d['provincia'],
                           'zona': zona, 'dias': dias_pendiente(d['fecha_entrada'])})
    return render_template('mapa.html', puntos=puntos)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
